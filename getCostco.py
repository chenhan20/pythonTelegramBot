import asyncio
import requests
import json
import prettytable as pt
import datetime
import telegramBot
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import getDb

url = 'https://www.costco.com.tw/rest/v2/taiwan/products/search?fields=FULL&query=&pageSize=100&category=hot-buys&lang=zh_TW&curr=TWD'
now = datetime.datetime.now()
file_name = 'costcoPrice.xlsx'

async def getData():
    response = requests.get(url)
    if response.status_code != 200:
        return []
    date_str = now.strftime("%Y%m%d")
    costco_data = json.loads(response.text)
    products = costco_data['products']
    data_list = []
    # 印出所有商品資訊
    for product in products:
        coupon = product.get('couponDiscount')
        if coupon is not None:
            try:
                data_list.append({
                    'title': product['name'],
                    'price': product['basePrice']['formattedValue'],
                    'discount': -float(coupon['discountValue']),
                    'result': '$' + str(product['basePrice']['value'] - coupon['discountValue'])
                })
            except Exception as e:
                print('error:', e)
                continue

    to_excel(data_list)
    # toTxt(data_list, date_str)
    telegram_ids = getDb.getCostcoTelegramIds()
    for sendId in telegram_ids:
        # 使用 with 確保檔案正確開啟和關閉
        with open(file_name, 'rb') as file:
            await telegramBot.sendFile(sendId, file,  f'{date_str} 好事多商品特價訊息\n 若不想收到請輸入\n/endCostco')


def toTxt(costcoData, date_str):
    tb1 = pt.PrettyTable()
    tb1.set_style(pt.PLAIN_COLUMNS)
    col1 = '名稱'
    col2 = '價格 | 折扣 | 結果'
    tb1.field_names = [col1, col2]
    tb1.align[col1] = "l"
    tb1.align[col2] = "l"

    title = f'{date_str} - COSTCO 特價商品'

    for data in costcoData:
        try:
            index_title = data['title']
            index_value = f"{data['price']} | {data['discount']} | {data['result']}"
            tb1.add_row([index_title, index_value])
        except Exception as error:
            print("Oops! An exception has occured:", error)
            print("Exception TYPE:", type(error))
    tbStr = f'{title}\n{tb1.get_string()}'
    with open("costcoPrice.txt", "w+", encoding='UTF-8') as f:
        f.write(tbStr)


def to_excel(costco_data):
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = 'costco'

    # 設定表頭
    headers = ['名稱', '價格', '折扣', '結果']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header).font = Font(bold=True)  # 粗體表頭

    # 設定內容
    text_max_length = 0
    for idx, data in enumerate(costco_data):
        excel_start_row = idx + 2  # 資料從第2行開始
        sheet['A' + str(excel_start_row)] = data['title']
        sheet['B' + str(excel_start_row)] = data['price']
        sheet['C' + str(excel_start_row)] = data['discount']
        sheet['D' + str(excel_start_row)] = data['result']
        # 設定右對齊
        sheet['B' + str(excel_start_row)].alignment = Alignment(horizontal='right')
        sheet['C' + str(excel_start_row)].alignment = Alignment(horizontal='right')
        sheet['D' + str(excel_start_row)].alignment = Alignment(horizontal='right')

        # 調整名稱欄位寬度
        if len(data['title']) > text_max_length:
            text_max_length = len(data['title'])
        sheet.column_dimensions['A'].width = text_max_length * 1.5
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 10

        # 將負數設置為紅色
        for col in ['B', 'C', 'D']:
            cell = sheet[col + str(excel_start_row)]
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = Font(color="FF0000")  # 紅色字體
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 紅色背景
    # 儲存檔案
    workbook.save(file_name)


if __name__ == '__main__':
    asyncio.run(getData())
